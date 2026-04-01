import customtkinter as ctk
from tkinter import messagebox
import webbrowser
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

# =============================================================================
# --- LÓGICA DE EXCEL (Antigo tabela_auto.py) ---
# =============================================================================

def estilizar(sheet, lista, estilo, tipo):
    """
    lista: ['A1', 'B1']
    estilo: o objeto de estilo (Font, PatternFill, etc.)
    tipo: string com o nome da propriedade ('font', 'fill', 'alignment', 'border')
    """
    for ref in lista:
        setattr(sheet[ref], tipo, estilo)

def header(ano, mes, moeda, smp):
    wb = openpyxl.Workbook()
    wb.create_sheet(f'Tabela Custo de Entrada&Saida-{mes}_{ano}')
    wb.remove(wb['Sheet'])

    sheet = wb[f'Tabela Custo de Entrada&Saida-{mes}_{ano}']
    sheet['F1'] = 'Ano:'
    sheet['F2'] = 'Mês:'
    sheet['F3'] = 'Moeda:'

    sheet['G1'] = ano
    sheet['G2'] = mes
    sheet['G3'] = moeda

    sheet['A5'] = 'DIÁRIO DE CAIXA'
    sheet.merge_cells('A5:G5')

    sheet['A6'] = 'Nº'
    sheet.merge_cells('A6:A7')
    sheet['B6'] = 'DATA'
    sheet.merge_cells('B6:B7')
    sheet['C6'] = 'DESIGNAÇÃO'
    sheet.merge_cells('C6:D7')
    sheet['E6'] = 'ENTRADAS (+)'
    sheet.merge_cells('E6:E7')
    sheet['F6'] = 'SAÍDAS (-)'
    sheet.merge_cells('F6:F7')
    sheet['G6'] = 'SALDO'
    sheet.merge_cells('G6:G7')

    sheet['D9'] = 'Saldo do Mês Anterior'
    sheet['G9'] = ' ' + str(smp)

    celula = sheet['A5']
    celula.alignment = Alignment(horizontal='center', vertical='center')

    for col in ['A6', 'C6', 'B6', 'E6', 'F6', 'G6']:
        sheet[col].alignment = Alignment(horizontal='center', vertical='center')

    for cell in ['F1', 'F2', 'F3', 'G1', 'G2', 'G3']:
        sheet[cell].font = Font(name='Times New Roman', size=11, bold=True, color="000000")

    sheet['A5'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")

    for cell in ['A6', 'B6', 'C6', 'E6', 'F6', 'G6', 'D9', 'E9', 'G9']:
        sheet[cell].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")

    branco = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
    cinzento = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    laranja = PatternFill(fill_type='solid', start_color='ffc000', end_color='ffc000')

    brancos = ['A1', 'A2', 'A3', 'A4', 'A5', 'B1', 'B2', 'B3', 'B4', 'C1', 'C2', 'C3', 'C4',
                'D1', 'D2', 'D3', 'D4', 'E1', 'E2', 'E3', 'E4', 'F1', 'F2', 'F3', 'F4',
                'G1', 'G2', 'G3', 'G4']
    cinzentos = ['A6', 'B6', 'C6', 'D6', 'D9', 'E6', 'E9', 'F6', 'F9', 'G6', 'G9']
    laranjas = ['A8', 'A9', 'B8', 'B9', 'C8', 'C9', 'D8', 'E8', 'F8', 'G8']

    estilizar(sheet, brancos, branco, 'fill')
    estilizar(sheet, cinzentos, cinzento, 'fill')
    estilizar(sheet, laranjas, laranja, 'fill')

    lista2 = ['D9', 'E9', 'F9', 'G9']
    lista3 = ['A6', 'B6', 'C6', 'E6', 'F6', 'G6', 'A7', 'B7', 'C7', 'E7', 'F7', 'G7']

    linha_fina = openpyxl.styles.Side(color='000000', style='thin')
    linha_media = openpyxl.styles.Side(color='000000', style='medium')
    linha_grossa = openpyxl.styles.Side(color='000000', style='thick')

    bordas_finas = openpyxl.styles.Border(left=linha_fina, right=linha_fina, top=linha_fina, bottom=linha_fina)
    bordas_medias = openpyxl.styles.Border(left=linha_media, right=linha_media, top=linha_media, bottom=linha_media)

    estilizar(sheet, ['G1', 'G2', 'G3'], bordas_finas, 'border')
    estilizar(sheet, lista3, bordas_medias, 'border')
    estilizar(sheet, lista2, openpyxl.styles.Border(top=linha_media), 'border')
    estilizar(sheet, ['D9'], openpyxl.styles.Border(left=linha_media, top=linha_media), 'border')
    estilizar(sheet, ['G9'], openpyxl.styles.Border(left=linha_media, right=linha_media, top=linha_media), 'border')

    wb.save('Custo_Entrada&Saida.xlsx')

def body(lista, mes, ano):
    wb = openpyxl.load_workbook('Custo_Entrada&Saida.xlsx')    
    sheet = wb[f'Tabela Custo de Entrada&Saida-{mes}_{ano}']
    cont_table = 10 

    for i in range(len(lista)):
        item = lista[i]
        sheet.merge_cells(f'C{cont_table}:D{cont_table}') 

        sheet[f'A{cont_table}'] = item[0]
        sheet[f'B{cont_table}'] = item[1]
        sheet[f'C{cont_table}'] = item[2]
        sheet.alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'E{cont_table}'] = item[3]
        sheet[f'F{cont_table}'] = item[4]
        sheet[f'G{cont_table}'] = item[5]

        for col in ['A', 'B', 'C', 'E', 'F', 'G']:
            sheet[f'{col}{cont_table}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        
        linha_fina = openpyxl.styles.Side(color='000000', style='thin')
        bordas_finas = openpyxl.styles.Border(left=linha_fina, right=linha_fina, top=linha_fina, bottom=linha_fina)
        
        for col in ['A', 'B', 'C', 'E', 'F', 'G']:
            sheet[f'{col}{cont_table}'].border = bordas_finas
            
        cont_table += 1
    wb.save('Custo_Entrada&Saida.xlsx')

def footer(lista, mes, ano, sexo, nome, saldo_atual, p_total, saldo_anterior, despesa_anterior, ativos_total):
    wb = openpyxl.load_workbook('Custo_Entrada&Saida.xlsx')    
    sheet = wb[f'Tabela Custo de Entrada&Saida-{mes}_{ano}']

    tamanho = len(lista) + 10 

    branco = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
    cinzento = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    laranja = PatternFill(fill_type='solid', start_color='ffc000', end_color='ffc000')

    lista1 = [f'A{tamanho}', f'B{tamanho}', f'C{tamanho}', f'D{tamanho}', f'F{tamanho}', f'E{tamanho + 2}', f'F{tamanho + 2}', f'G{tamanho + 2}', f'G{tamanho + 4}']
    lista2 = [f'E{tamanho + 3}', f'F{tamanho + 3}', f'G{tamanho + 3}']

    estilizar(sheet, lista1, cinzento, 'fill')
    estilizar(sheet, [f'E{tamanho}', f'G{tamanho}'], branco, 'fill')
    estilizar(sheet, [f'A{tamanho + 1}', f'B{tamanho + 1}', f'C{tamanho + 1}', f'D{tamanho + 1}', f'E{tamanho + 1}', f'F{tamanho + 1}', f'G{tamanho + 1}'], branco, 'fill')
    estilizar(sheet, [f'D{tamanho + 2}', f'D{tamanho + 3}', f'D{tamanho + 4}', f'A{tamanho + 5}', f'A{tamanho + 6}', f'A{tamanho + 7}', f'B{tamanho + 5}', f'B{tamanho + 6}', f'B{tamanho + 7}', f'C{tamanho + 5}', f'C{tamanho + 6}', f'C{tamanho + 7}'], branco, 'fill')

    linha_fina = openpyxl.styles.Side(color='000000', style='thin')
    linha_media = openpyxl.styles.Side(color='000000', style='medium')
    bordas_medias = openpyxl.styles.Border(right=linha_media, left=linha_media, top=linha_media, bottom=linha_media)

    estilizar(sheet, lista2, laranja, 'fill')
    estilizar(sheet, lista1, bordas_medias, 'border')
    estilizar(sheet, lista2, bordas_medias, 'border')
    estilizar(sheet, [f'E{tamanho}', f'A{tamanho + 7}', f'B{tamanho + 7}', f'C{tamanho + 7}'], openpyxl.styles.Border(bottom=linha_fina), 'border')
    estilizar(sheet, [f'G{tamanho}'], openpyxl.styles.Border(right=linha_fina, bottom=linha_fina), 'border')

    sheet[f'D{tamanho + 2}'] = 'MOVIMENTOS DO MÊS'
    sheet[f'D{tamanho + 3}'] = 'SALDO DO MÊS ANTERIOR'
    sheet[f'D{tamanho + 4}'] = 'SALDO DO MÊS SEGUINTE'
    sheet[f'E{tamanho + 4}'] = '|------------»»»»»»»»»»»»'

    sheet[f'F{tamanho}'] = float(p_total)
    sheet[f'G{tamanho}'] = float(saldo_atual)
    sheet[f'E{tamanho + 2}'] = float(ativos_total)
    sheet[f'F{tamanho + 2}'] = float(p_total)
    sheet[f'G{tamanho + 2}'] = float(ativos_total - p_total)

    sheet[f'E{tamanho + 3}'] = float(saldo_anterior)
    sheet[f'F{tamanho + 3}'] = float(despesa_anterior)
    sheet[f'G{tamanho + 3}'] = float(saldo_anterior - despesa_anterior)

    sheet[f'G{tamanho + 4}'] = '+' + str(float(saldo_anterior - despesa_anterior + saldo_atual))

    for cell in [f'E{tamanho + 2}', f'F{tamanho + 2}', f'G{tamanho + 2}', f'G{tamanho + 4}', f'F{tamanho}']:
        sheet[cell].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    
    for cell in [f'E{tamanho + 3}', f'F{tamanho + 3}', f'G{tamanho + 3}', f'G{tamanho}']:
        sheet[cell].font = Font(name='Times New Roman', size=11, bold=True, color="000000")

    sheet.merge_cells(f'A{tamanho}:D{tamanho}')
    sheet.merge_cells(f'A{tamanho + 5}:B{tamanho + 5}')
    sheet.merge_cells(f'A{tamanho + 6}:B{tamanho + 6}')

    if sexo.upper() == 'F':
        sheet[f'A{tamanho + 5}'] = 'A TESOUREIRA'
    elif sexo.upper() == 'M':
        sheet[f'A{tamanho + 5}'] = 'O TESOUREIRO'
    else:
        sheet[f'A{tamanho + 5}'] = 'O/A TESOUREIRO/A'

    sheet[f'A{tamanho + 6}'] = str(nome)

    for cell in [f'D{tamanho + 2}', f'D{tamanho + 3}', f'D{tamanho + 4}']:
        sheet[cell].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        
    for cell in [f'A{tamanho + 5}', f'A{tamanho + 6}']:
        sheet[cell].font = Font(name='Times New Roman', size=11, bold=False, color="000000")

    wb.save('Custo_Entrada&Saida.xlsx')

# =============================================================================
# --- CONFIGURAÇÃO GLOBAL E GUI (Antigo app.py) ---
# =============================================================================

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

lista_itens = []
widgets_edicao = {}

def limpar_palco():
    """Remove todos os widgets do frame principal para trocar de tela."""
    for widget in palco_conteudo.winfo_children():
        widget.destroy()

def mostrar_sobre():
    """Desenha a tela 'Sobre' com informações e guia de instalação."""
    limpar_palco()

    ctk.CTkLabel(palco_conteudo, text="Sobre o Sistema", font=("Arial", 24, "bold")).pack(pady=(20, 10))
    
    frame_info = ctk.CTkFrame(palco_conteudo)
    frame_info.pack(fill="x", padx=30, pady=10)
    
    ctk.CTkLabel(frame_info, text="Desenvolvedor: Cristiano Glória", font=("Arial", 14, "bold")).pack(pady=5)
    ctk.CTkLabel(frame_info, text="Suporte Técnico: cristianoakgloria@gmail.com", font=("Arial", 12)).pack()
    
    link_github = ctk.CTkLabel(frame_info, text="Código Fonte: github.com/cristianoakgloria-web/Automatizacao...", 
                               text_color="#1f538d", cursor="hand2", font=("Arial", 12, "underline"))
    link_github.pack(pady=5)
    link_github.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/cristianoakgloria-web/Automatizacao-de-Tabela-de-Custo-de-Entrada-e-Saida"))

    ctk.CTkLabel(palco_conteudo, text="Instruções de Uso:", font=("Arial", 16, "bold")).pack(pady=(15, 5))
    txt_uso = ctk.CTkTextbox(palco_conteudo, height=100, width=600)
    txt_uso.pack(padx=30, pady=5)
    txt_uso.insert("0.0", "1. Clique em 'Nova Tabela' e preencha os dados de configuração.\n"
                          "2. Adicione os itens. Os campos limpam-se automaticamente após adicionar.\n"
                          "3. Se precisar editar, altere os valores na lista e clique em 'Salvar' na linha correspondente.\n"
                          "4. Clique em 'GERAR EXCEL FINAL' para criar o ficheiro .xlsx.")
    txt_uso.configure(state="disabled")

    ctk.CTkLabel(palco_conteudo, text="Como Gerar Executável (PyInstaller):", font=("Arial", 16, "bold")).pack(pady=(15, 5))
    txt_install = ctk.CTkTextbox(palco_conteudo, height=130, width=600, fg_color="black", text_color="#00FF00")
    txt_install.pack(padx=30, pady=5)
    
    comando_pyinst = (
        "Use o PyInstaller no terminal para Mac, Windows ou Linux:\n\n"
        "1. Instalar: pip install pyinstaller\n"
        "2. Gerar: pyinstaller --noconsole --onefile --collect-all customtkinter app.py\n\n"
        "O executável será criado na pasta 'dist'."
    )
    txt_install.insert("0.0", comando_pyinst)
    txt_install.configure(state="disabled")

def mostrar_novaTabela():
    """Desenha a interface de criação de tabela."""
    global widgets_edicao
    widgets_edicao = {}
    limpar_palco()

    ctk.CTkLabel(palco_conteudo, text="1. Configuração do Diário", font=("Arial", 16, "bold")).pack(pady=(10, 5))
    frame_config = ctk.CTkFrame(palco_conteudo)
    frame_config.pack(fill="x", padx=20, pady=5)

    ent_ano = ctk.CTkEntry(frame_config, placeholder_text="Ano (ex: 2026)"); ent_ano.grid(row=0, column=0, padx=5, pady=5)
    ent_mes = ctk.CTkEntry(frame_config, placeholder_text="Mês (ex: MARÇO)"); ent_mes.grid(row=0, column=1, padx=5, pady=5)
    ent_moeda = ctk.CTkEntry(frame_config, placeholder_text="Moeda (ex: AKZ)"); ent_moeda.grid(row=0, column=2, padx=5, pady=5)
    ent_smp = ctk.CTkEntry(frame_config, placeholder_text="Saldo Mês Anterior"); ent_smp.grid(row=1, column=0, padx=5, pady=5)
    ent_des_ant = ctk.CTkEntry(frame_config, placeholder_text="Despesa Mês Anterior"); ent_des_ant.grid(row=1, column=1, padx=5, pady=5)

    ctk.CTkLabel(palco_conteudo, text="2. Adicionar Movimentação", font=("Arial", 14, "bold")).pack(pady=(10, 5))
    frame_add = ctk.CTkFrame(palco_conteudo)
    frame_add.pack(fill="x", padx=20, pady=5)

    e_data = ctk.CTkEntry(frame_add, placeholder_text="Data (dd-mm)"); e_data.grid(row=0, column=0, padx=5, pady=5)
    e_desc = ctk.CTkEntry(frame_add, placeholder_text="Designação"); e_desc.grid(row=0, column=1, padx=5, pady=5)
    e_ent = ctk.CTkEntry(frame_add, placeholder_text="Valor Entrada"); e_ent.grid(row=1, column=0, padx=5, pady=5)
    e_sai = ctk.CTkEntry(frame_add, placeholder_text="Valor Saída"); e_sai.grid(row=1, column=1, padx=5, pady=5)

    scroll_frame = ctk.CTkScrollableFrame(palco_conteudo, height=180)
    scroll_frame.pack(fill="both", expand=True, padx=20, pady=10)

    def atualizar_tabela_visual():
        for w in scroll_frame.winfo_children(): 
            w.destroy()
        
        headers = ["Nº", "Data", "Designação", "Entrada", "Saída", "Ações"]
        for i, h in enumerate(headers): 
            ctk.CTkLabel(scroll_frame, text=h, font=("Arial", 11, "bold")).grid(row=0, column=i, padx=10)

        for idx, item in enumerate(lista_itens):
            row = idx + 1
            ctk.CTkLabel(scroll_frame, text=str(item[0])).grid(row=row, column=0)
            
            ed_dat = ctk.CTkEntry(scroll_frame, width=80); ed_dat.insert(0, item[1]); ed_dat.grid(row=row, column=1, padx=2)
            ed_des = ctk.CTkEntry(scroll_frame, width=150); ed_des.insert(0, item[2]); ed_des.grid(row=row, column=2, padx=2)
            ed_ent = ctk.CTkEntry(scroll_frame, width=80); ed_ent.insert(0, str(item[3])); ed_ent.grid(row=row, column=3, padx=2)
            ed_sai = ctk.CTkEntry(scroll_frame, width=80); ed_sai.insert(0, str(item[4])); ed_sai.grid(row=row, column=4, padx=2)
            
            widgets_edicao[idx] = [ed_dat, ed_des, ed_ent, ed_sai]
            ctk.CTkButton(scroll_frame, text="Salvar", width=60, fg_color="gray", 
                          command=lambda i=idx: salvar_edicao_linha(i)).grid(row=row, column=5, padx=2)

    def salvar_edicao_linha(i):
        try:
            w = widgets_edicao[i]
            lista_itens[i][1] = w[0].get()
            lista_itens[i][2] = w[1].get()
            lista_itens[i][3] = float(w[2].get() or 0)
            lista_itens[i][4] = float(w[3].get() or 0)
            recalcular_tudo()
            messagebox.showinfo("Sucesso", f"Linha {i+1} atualizada e saldos recalculados.")
        except ValueError:
            messagebox.showerror("Erro", "Valores numéricos inválidos na edição.")

    def recalcular_tudo():
        try:
            saldo_anterior = float(ent_smp.get() or 0) - float(ent_des_ant.get() or 0)
            for it in lista_itens:
                it[5] = saldo_anterior + it[3] - it[4]
                saldo_anterior = it[5]
            atualizar_tabela_visual()
        except ValueError: pass

    def adicionar_item():
        try:
            s_ini = float(ent_smp.get() or 0) - float(ent_des_ant.get() or 0)
            s_acumulado = lista_itens[-1][5] if lista_itens else s_ini
            v_ent, v_sai = float(e_ent.get() or 0), float(e_sai.get() or 0)
            
            novo_saldo = s_acumulado + v_ent - v_sai
            lista_itens.append([len(lista_itens)+1, e_data.get(), e_desc.get(), v_ent, v_sai, novo_saldo])
            
            e_data.delete(0, 'end'); e_desc.delete(0, 'end'); e_ent.delete(0, 'end'); e_sai.delete(0, 'end')
            
            atualizar_tabela_visual()
        except ValueError:
            messagebox.showerror("Erro", "Por favor, insira números válidos nos valores.")

    ctk.CTkButton(palco_conteudo, text="Adicionar Item à Lista", command=adicionar_item).pack(pady=5)

    frame_fim = ctk.CTkFrame(palco_conteudo)
    frame_fim.pack(fill="x", padx=20, pady=10)
    e_nome = ctk.CTkEntry(frame_fim, placeholder_text="Nome do Tesoureiro/a", width=200); e_nome.grid(row=0, column=0, padx=5)
    e_sexo = ctk.CTkEntry(frame_fim, placeholder_text="M/F", width=60); e_sexo.grid(row=0, column=1, padx=5)

    def finalizar_e_gerar():
        if not lista_itens:
            messagebox.showwarning("Aviso", "A lista está vazia.")
            return
        try:
            recalcular_tudo()
            smp_val = float(ent_smp.get() or 0)
            des_val = float(ent_des_ant.get() or 0)
            
            header(ent_ano.get(), ent_mes.get().upper(), ent_moeda.get().upper(), float(smp_val - des_val))
            body(lista_itens, ent_mes.get().upper(), ent_ano.get())
            footer(lista_itens, ent_mes.get().upper(), ent_ano.get(), e_sexo.get(), e_nome.get(), 
                   lista_itens[-1][5], sum(x[4] for x in lista_itens), smp_val, des_val, sum(x[3] for x in lista_itens))
            
            messagebox.showinfo("Sucesso", f"O ficheiro 'Custo_Entrada&Saida.xlsx' foi gerado para {ent_mes.get()}!")
        except Exception as err:
            messagebox.showerror("Erro Crítico", f"Falha ao gerar Excel: {err}")

    btn_gerar = ctk.CTkButton(palco_conteudo, text="GERAR EXCEL FINAL", fg_color="green", 
                               hover_color="darkgreen", font=("Arial", 14, "bold"), command=finalizar_e_gerar)
    btn_gerar.pack(pady=10)
    
    atualizar_tabela_visual()

# =============================================================================
# --- INICIALIZAÇÃO DA JANELA PRINCIPAL ---
# =============================================================================

if __name__ == "__main__":
    app = ctk.CTk()
    app.title("Sistema de Automatização de Tabelas - Cristiano Glória")
    app.geometry("1000x800")

    app.grid_columnconfigure(1, weight=1)
    app.grid_rowconfigure(0, weight=1)

    sidebar = ctk.CTkFrame(app, width=200, corner_radius=0)
    sidebar.grid(row=0, column=0, sticky="nsew")

    ctk.CTkLabel(sidebar, text="MENU", font=("Arial", 20, "bold")).pack(pady=30)
    ctk.CTkButton(sidebar, text="Nova Tabela", command=mostrar_novaTabela).pack(pady=10, padx=15)
    ctk.CTkButton(sidebar, text="Sobre o App", command=mostrar_sobre).pack(pady=10, padx=15)

    palco_conteudo = ctk.CTkFrame(app, corner_radius=15)
    palco_conteudo.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")

    mostrar_novaTabela()

    app.mainloop()
